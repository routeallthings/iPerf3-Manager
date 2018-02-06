#!/usr/bin/env python
'''
---AUTHOR---
Name: Matt Cross
Email: routeallthings@gmail.com

---PREREQ---
INSTALL iperf3 (pip install iperf3)

'''
'''Module Imports (Native)'''
import os
import sys
import time
import threading

## XLSX Needed Modules ##
try:
	from openpyxl import load_workbook
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('openpyxl module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install openpyxl')
		from openpyxl import load_workbook
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of Pandas. Please install manually and retry'
		sys.exit()
#
try:
	import fileinput
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('FileInput module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install FileInput')
		import FileInput
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of FileInput. Please install manually and retry'
		sys.exit()
# Darth-Veitcher Module https://github.com/darth-veitcher/xlhelper
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import OrderedDict
try:
	import xlhelper
except ImportError:
	requestsinstallstatus = fullpath = raw_input ('xlhelper module is missing, would you like to automatically install? (Y/N): ')
	if 'Y' in requestsinstallstatus or 'y' in requestsinstallstatus or 'yes' in requestsinstallstatus or 'Yes' in requestsinstallstatus or 'YES' in requestsinstallstatus:
		os.system('python -m pip install git+git://github.com/routeallthings/xlhelper.git')
		import xlhelper
	else:
		print 'You selected an option other than yes. Please be aware that this script requires the use of xlhelper. Please install manually and retry'
		sys.exit()

## iPerf3 Module ##
try:
	import iperf3
except ImportError:
	iperf3installstatus = fullpath = raw_input ('iperf3 module is missing, would you like to automatically install? (Y/N): ')
	if "Y" in iperf3installstatus.upper() or "YES" in iperf3installstatus.upper():
		try:
			os.system('python -m pip install iperf3')
			import iperf3
		except ImportError:
			if 'win32' in PlatformType:
				print 'could not find pip, please reinstall python for windows'
			if 'linux' in PlatformType:
				try:
					os.system('sudo yum install python-pip -y')
					os.system('sudo apt-get install python-pip -y')
				except:
					print 'Error installing pip. Please install pip and rerun the script'
					sys.exit()
				try: 
					os.system('python -m pip install iperf3')
					import iperf3
				except:
					print 'Error installing the iperf3 module, please install manually'
					sys.exit()
	else:		
		print "You selected an option other than yes. Please be aware that this script requires the use of iperf3. Please install manually and retry"
		sys.exit()

# Get Platform Type
PlatformType = sys.platform
	
##### Configuration Parameters for NON-XLSX mode #####
Modeq = 'Server' # Specify Client or Server
Durationq = 30 # Specify a length in seconds for the test
Serverq = '10.10.15.10' # Specify either the remote server (Client Mode) or the IP of the Bind Address (Server Mode)
Portq = 5201 # Specify the port or leave as is for the default
Protocolq = 'udp' # Specify udp or tcp
QuietModeq = 'True' # Specify (True/False) if you want it to run in the screen
Bandwidthq = 1000000000 # If the protocol is UDP, set the the bandwidth in bits per second
Reversemodeq = '' # (True/False) Specify the direction of the flow
		
##### Configuration Parameter for XLSX mode #####
XLSXMode = 'True' # (True/False) If true it will automatically pull configuration data from a XLSX template and override any NON-XLSX settings
XLSXLocation = '' # Put in the full path to the XLSX file, otherwise leave blank (nothing in between two quotes) for it to prompt


################# Start Of Script #################

#Functions#

try:
	iperf3test = iperf3.Client()

except OSError:
	try:
		print 'iperf3 was not detected. Attempting to install.'
		os.system('sudo yum install iperf3 -y')
		os.system('sudo apt-get install iperf3 -y')
	except:
		print 'Error installing iperf3. Please install iperf3 and rerun the script'
		sys.exit()

def IPERF3_nonXLSX():
	if 'client' in Modeq.lower():
		iperf3test = iperf3.Client()
		iperf3test.server_hostname = Serverq
		iperf3test.duration = Durationq
		iperf3test.protocol = Protocolq
		iperf3test.bandwidth = int(Bandwidthq)
		if 'true' in Reversemodeq.lower():
			iperf3test.reverse = 1
	if 'server' in Modeq.lower():
		iperf3test = iperf3.Server()
		iperf3test.bind_address = Serverq
	iperf3test.port = Portq
	if 'True' in QuietModeq:
		iperf3test.json_output = True
		iperf3test.verbose = False
	# Client Portion
	if 'client' in Modeq.lower():
		print('Connecting to {0}:{1}'.format(iperf3test.server_hostname, iperf3test.port))
		result = iperf3test.run()
		if result.error:
			print(result.error)
		else:
			print('')
			print('Test completed:')
			print('  started at         {0}'.format(result.time))
			print('  bytes transmitted  {0}'.format(result.bytes))
			print('  jitter (ms)        {0}'.format(result.jitter_ms))
			print('  avg cpu load       {0}%\n'.format(result.local_cpu_total))

			print('Average transmitted data in all sorts of networky formats:')
			print('  bits per second      (bps)   {0}'.format(result.bps))
			print('  Kilobits per second  (kbps)  {0}'.format(result.kbps))
			print('  Megabits per second  (Mbps)  {0}'.format(result.Mbps))
			print('  KiloBytes per second (kB/s)  {0}'.format(result.kB_s))
			print('  MegaBytes per second (MB/s)  {0}'.format(result.MB_s))
	# Server Portion
	if 'server' in Modeq.lower():
		while True:
			iperf3test.run()

def IPERF3_XLSX(rowdata):
	t = threading.currentThread()
	while getattr(t, "do_run", True) and runtime < maxruntime:
		sshdevicetype = 'linux'
		# Device Configuration
		try:
			remotedevice = rowdata.get('Remote Device').encode('utf-8')
		except:
			remotedevice = rowdata.get('Remote Device')
			remotedevice = str(remotedevice)
		try:
			Modeq = rowdata.get('Mode').encode('utf-8')
		except:
			Modeq = rowdata.get('Mode')
			Modeq = str(Modeq)
		try:
			Durationq = rowdata.get('Duration').encode('utf-8')
		except:
			Durationq = rowdata.get('Duration')
			Durationq = str(Durationq)
		try:
			Serverq = rowdata.get('Server').encode('utf-8')
		except:
			Serverq = rowdata.get('Server')
			Serverq = str(Serverq)
		try:
			Portq = rowdata.get('Port').encode('utf-8')
		except:
			Portq = rowdata.get('Port')
			Portq = str(Portq)
		try:
			Protocolq = rowdata.get('Protocol').encode('utf-8')
		except:
			Protocolq = rowdata.get('Protocol')
			Protocolq = str(Protocolq)
		try:
			Bandwidthq = rowdata.get('Target Bandwidth').encode('utf-8')
		except:
			Bandwidthq = rowdata.get('Target Bandwidth')
			Bandwidthq = str(Bandwidthq)
		try:
			Reversemodeq = rowdata.get('Reverse Mode').encode('utf-8')
		except:
			Reversemodeq = rowdata.get('Reverse Mode')
			Reversemodeq = str(Reversemodeq)
		#Start Connection
		try:
			sshnet_connect = ConnectHandler(device_type=sshdevicetype, ip=remotedevice, username=sshusername, password=sshpassword)
			sshdevicehostname = sshnet_connect.find_prompt()
			print 'Successfully connected to ' + remotedevice
			if 'client' in Modeq.lower():
				iperf3test = iperf3.Client()
				iperf3test.server_hostname = Serverq
				iperf3test.duration = Durationq
				iperf3test.protocol = Protocolq
				iperf3test.bandwidth = int(Bandwidthq)
				if 'true' in Reversemodeq.lower():
					iperf3test.reverse = 1
			if 'server' in Modeq.lower():
				iperf3test = iperf3.Server()
				iperf3test.bind_address = Serverq
			iperf3test.port = Portq
			# Client Portion
			if 'client' in Modeq.lower():
				print('Connecting to {0}:{1}'.format(iperf3test.server_hostname, iperf3test.port))
				result = iperf3test.run()
				writer.writerow({'Device IP': remotedevice, 'Start Time': iperf3test.time, 'Average CPU Load': iperf3test.local_cpu_total, 'Total Bytes Transmitted': iperf3test.bytes, 'Mbps': iperf3test.Mbps, 'Average Jitter': iperf3test.jitter_ms})
			# Server Portion
			if 'server' in Modeq.lower():
				while True:
					iperf3test.run()
			sshnet_connect.disconnect()
		except:
			print "Failure in connecting to device " + sshdevicename + ". Please confirm that the IP address is correct"
			sshnet_connect.disconnect()
	print("Stopping as you wish.")
	try:
		sshnet_connect.disconnect()
	except:
		''' Skip '''
# XLSX Mode Start #
if __name__ == "__main__":
	if 'true' in XLSXMode.lower():
		saveresultslist = []
		Modeq = ''
		Durationq = ''
		Serverq = ''
		Portq = ''
		Protocolq = ''
		Bandwidthq = ''
		Reversemodeq = ''
		QuietModeq = ''
		username = ''
		password = ''
		exportlocation = ''
		maxruntime = ''
		if XLSXLocation == '':
			XLSXLocation = 'C:/Python27/iperf3template.xlsx'
		for configdata in xlhelper.sheet_to_dict(XLSXLocation,'Config'):
			try:
				Variable = rowdata.get('Variable').encode('utf-8')
			except:
				Variable = rowdata.get('Variable')
				Variable = str(Variable)
			try:
				Value = rowdata.get('Value').encode('utf-8')
			except:
				Value = rowdata.get('Value')
				Value = str(Value)			
			if 'username' in Variable.lower():
				sshusername = Value
				if sshusername == '':
					sshusername = raw_input('What SSH username do you want to use to login to these devices?:')
			if 'password' in Variable.lower():
				sshpassword = Value
				if sshpassword == '':
					sshpassword = raw_input('What password do you want to use to login to these devices?:')
			if 'exportlocation' in Variable.lower():
				exportlocation = Value
				if exportlocation == '':
					exportlocation = raw_input('What location do you want to export the CSV data to (e.g. C:\\Test\\Results.csv)?:')
		while True:
			for rowdata in xlhelper.sheet_to_dict(XLSXLocation,'Data'):
				try:
					remotedevice = rowdata.get('Remote Device').encode('utf-8')
				except:
					remotedevice = rowdata.get('Remote Device')
					remotedevice = str(remotedevice)
				#Parallel Processing
				print "Spawning Thread for " + remotedevice
				t = threading.Thread(target=IPERF3_XLSX, args=(rowdata,))
				t.start()
				main_thread = threading.currentThread()
			print 'All Threads Started'
			print 'Please wait for duration of the tests before selecting an option below'
			print''
			print '##############################################'
			print '          		  MENU'
			print '##############################################'
			print ''
			print '1. Cancel All Tests'
			print '2. Restart All Tests (Will Overwrite Results)'
			print '3. Exit'
			menuq = raw_input('Which option would you like (1-3)?:')
			if int(menuq) == 1:
				t.do_run = False
				print 'Cancelling in 5 seconds'
				time.sleep(5)
				break
			if int(menuq) == 2:
				t.do_run = False
				print 'Restarting tests in 5 seconds'
				time.sleep(5)
				continue
			if int(menuq) == 3:
				t.do_run = False
				print 'Exiting in 5 seconds'
				time.sleep(5)
				sys.exit()
	
# NON-XLSX Mode Start
if 'false' in XLSXMode.lower():
	IPERF3_nonXLSX()

# End of Script
print 'Script is complete'